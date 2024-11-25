# main.py

import os
from login import login, logout, register
from file_handler import cargar_archivo, listar_archivos
from data_processor import procesar_datos
from data_visualization import graficar_tendencias_historicas
from utils import aplicar_estilos
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import re

aplicar_estilos()

st.markdown("""
    <style>
    @import url('styles.css');
    </style>
""", unsafe_allow_html=True)

st.title("Sistema de Visualización y Consolidación de Datos SNIES")

@st.cache_data
def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    writer.save()
    processed_data = output.getvalue()
    return processed_data

@st.cache_data
def listar_archivos_por_anio(anio_inicio, anio_fin):
    input_dir = "docs/inputs"
    archivos = [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f))]
    archivos_relevantes = []

    pattern = re.compile(r'(\d{4})')
    for archivo in archivos:
        match = pattern.search(archivo)
        if match:
            year = int(match.group(0))
            if anio_inicio <= year <= anio_fin:
                archivo_path = os.path.join(input_dir, archivo)
                # Cargar el archivo en memoria
                with open(archivo_path, 'rb') as f:
                    in_memory_file = BytesIO(f.read())
                    in_memory_file.name = archivo  # Añadir nombre para similitud con st.file_uploader
                    archivos_relevantes.append(in_memory_file)

    return archivos_relevantes

if 'user' not in st.session_state:
    option = st.sidebar.selectbox("Seleccione una opción", ["Iniciar Sesión", "Registrarse"])
    if option == "Iniciar Sesión":
        user = login()
    else:
        user = register()

    if user:
        st.session_state['user'] = user
else:
    st.write(f"Bienvenido, {st.session_state['user']['email']}")
    logout()

    st.subheader("Carga de Información")
    input_files = listar_archivos()

    st.write(f"Archivos disponibles ({len(input_files)} archivos):")
    st.write(input_files)

    tab1, tab2 = st.tabs(["Filtrar por Años y Programas", "Procesar y Visualizar Datos"])

    with tab1:
        st.title("Visualización de Estudiantes por Programa Académico en Excel usando Streamlit")


        def to_excel_with_style(df):
            output = BytesIO()
            wb = Workbook()
            ws = wb.active

            # Convertir DataFrame a filas de Excel
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)

            # Definir el rango de la tabla
            min_col = ws.min_column
            max_col = ws.max_column
            min_row = ws.min_row
            max_row = ws.max_row
            ref = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"

            # Crear una tabla en el rango de datos
            tab = Table(displayName="TablaResultados", ref=ref)

            # Añadir estilo a la tabla
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Guardar el archivo en memoria
            wb.save(output)
            processed_data = output.getvalue()
            return processed_data
        # Cargar los archivos Excel
        uploaded_files = st.file_uploader("Cargar archivos Excel", type="xlsx", accept_multiple_files=True)
        programa_academico = st.text_input("Ingrese el nombre del programa académico")

        if uploaded_files and programa_academico:
            resultados = procesar_datos(uploaded_files, programa_academico)

            # Mostrar la tabla en Streamlit
            st.write("Resultados del análisis:")
            st.dataframe(resultados)
            st.subheader("Descargar resultados")
            df_resultados = pd.DataFrame(resultados)

            st.download_button(
                label="Descargar Excel con formato",
                data=to_excel_with_style(resultados),
                file_name='resultados_con_formato.xlsx'
            )

            st.download_button(
                label="Descargar JSON",
                data=df_resultados.to_json(orient='records'),
                file_name='resultados.json',
                mime='application/json'
            )

            st.download_button(
                label="Descargar CSV",
                data=df_resultados.to_csv(index=False).encode('utf-8'),
                file_name='resultados.csv'
            )

            st.subheader("Visualización de Información")




